import smtplib
import string
import secrets
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from datetime import datetime

load_dotenv()

EXCEL_FILE = "codigo_descuento_clientes.xlsx"

# 1. Generate a random alphanumeric string
def generate_token(length=6):
    chars = string.ascii_letters.upper() #+ string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))

# 2. Send email
def send_email(to_email, token):
    from_email = os.environ["FROM_EMAIL"]
    password = os.environ["EMAIL_PASSWORD"]

    # Create HTML message
    html_content = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    padding: 20px;
                    color: #333;
                }}
                .container {{
                    max-width: 600px;
                    margin: auto;
                    background-color: #ffffff;
                    padding: 30px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .footer {{
                    text-align: center;
                    margin-top: 20px;
                }}
                .logo {{
                    width: 25%;
                    height: auto;
                }}
                .token {{
                    background-color: #007bff;
                    color: white;
                    padding: 10px 15px;
                    font-size: 18px;
                    border-radius: 5px;
                    display: inline-block;
                    margin-top: 20px;
                    letter-spacing: 1px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>¡Hola!</h2>
                <p>Con el siguiente código podrás acceder a un descuento:</p>
                <p style="text-align: center;">
                    <span class="token">{token}</span>
                </p>
                <p>Recuerda que tienes 24 horas a partir del momento en que recibiste este correo.</p>
                <div class="footer">
                    <img src="https://englifeinstitute.github.io/logo/logo.png" alt="Logo" class="logo" />
                </div>
            </div>
        </body>
        </html>
    """

    # MIME setup
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "Tu código de descuento"
    msg["From"] = from_email
    msg["To"] = to_email

    # Add HTML part
    msg.attach(MIMEText(html_content, "html"))

    # Send via SMTP
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(from_email, password)
        server.send_message(msg)

# 3. Write email and token to Excel
def write_to_excel(email, token):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Email", "Token", "DateTime"])

    ws.append([email, token, datetime.now().isoformat()])
    wb.save(EXCEL_FILE)

# 4. Main process
def process_email(email):
    token = generate_token()
    print(f"Generated token for {email}: {token}")
    send_email(email, token)
    write_to_excel(email, token)
    return {email: token}

# Example usage
if __name__ == "__main__":
    email_input = input("Enter your email: ")
    result = process_email(email_input)
    print("Email sent. Token saved:", result)