import string
import secrets
import os
from openpyxl import Workbook, load_workbook
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables from .env file
load_dotenv()

SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
FROM_EMAIL = os.getenv("FROM_EMAIL")  # Your verified sender email

# 1. Generate a random alphanumeric string
def generate_token(length=6):
    chars = string.ascii_letters.upper() #+ string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))

# 2. Send email via SendGrid
def send_email(to_email, token):
    message = Mail(
        from_email=FROM_EMAIL,
        to_emails=to_email,
        subject='Your verification token',
        html_content = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    padding: 20px;
                    color: #333;
                }}
                .container {{
                    max-width: 600px;
                    margin: auto;
                    background-color: #ffffff;
                    padding: 30px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .footer {{
                    text-align: center;
                    margin-top: 20px;
                }}
                .logo {{
                    width: 25%;
                    height: auto;
                }}
                .token {{
                    background-color: #007bff;
                    color: white;
                    padding: 10px 15px;
                    font-size: 18px;
                    border-radius: 5px;
                    display: inline-block;
                    margin-top: 20px;
                    letter-spacing: 1px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>¡Hola!</h2>
                <p>Con el siguiente código podrás acceder a un descuento:</p>
                <p style="text-align: center;">
                    <span class="token">{token}</span>
                </p>
                <p>Recuerda que tienes 24 horas a partir del momento en que recibiste este correo.</p>
                <div class="footer">
                    <img src="https://englifeinstitute.github.io/logo/logo.png" alt="Logo" class="logo" />
                </div>
            </div>
        </body>
        </html>
    """
        )
    try:
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        print(f"Email sent to {to_email}. Status code: {response.status_code}")
    except Exception as e:
        print("Error sending email:", e)

# 3. Save email, token, and datetime to an Excel file using openpyxl
def save_to_excel(email, token):
    # Get current datetime
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Define the Excel file path
    file_path = "codigo_descuento_clientes.xlsx"

    
    # Check if the Excel file exists
    if os.path.exists(file_path):
        # Load the existing workbook and select the active sheet
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        # Create a new workbook and add the headers
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Email", "Token", "Datetime"])  # Add headers

    # Append the new data row
    sheet.append([email, token, current_datetime])

    # Save the workbook
    wb.save(file_path)
    print(f"Data saved to {file_path}")

# 4. Main function
def process_email(email):
    token = generate_token()
    print(f"Generated token for {email}: {token}")
    send_email(email, token)
    save_to_excel(email, token)  # Save email and token to Excel
    return {email: token}

# Example usage
if __name__ == "__main__":
    email_input = input("Enter your email: ")
    result = process_email(email_input)
    print("Email sent. Token saved:", result)